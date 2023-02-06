import openpyxl
from tkinter import *

def write_to_excel():
    # get the values from the input fields
    name = name_entry.get()
    age = age_entry.get()

    # open the Excel sheet
    workbook = openpyxl.load_workbook('./ext/chatGPT/data.xlsx')
    worksheet = workbook['Sheet1']

    # write the values to the next empty row
    next_row = worksheet.max_row + 1
    worksheet.cell(row=next_row, column=1).value = name
    worksheet.cell(row=next_row, column=2).value = age

    # save the changes to the Excel sheet
    workbook.save('./ext/chatGPT/data.xlsx')
    print("Values written to Excel sheet.")

# create the Tkinter window
root = Tk()
root.config(bg='black')

# create the widgets
name_label = Label(root, text="Name:", bg='black', fg='white')
name_entry = Entry(root)
age_label = Label(root, text="Age:", bg='black', fg='white')
age_entry = Entry(root)
submit_button = Button(root, text="Submit", command=write_to_excel, bg='gray')

# place the widgets in the window
name_label.grid(row=0, column=0)
name_entry.grid(row=0, column=1)
age_label.grid(row=1, column=0)
age_entry.grid(row=1, column=1)
submit_button.grid(row=2, column=1)

# start the event loop
root.mainloop()
